import  openpyxl  as  op  

#파일명
wb = op.load_workbook("test2.xlsx") 
#시트명
ws = wb["직원명부"] 

#"B1" Cell에 입력하기 - Sheet Cell 이용
ws.cell(row=1, column=2).value = "입력테스트1"
#"C1" Cell에 입력하기 - 엑셀 인덱스(Range) 이용
ws["C1"].value = "입력테스트2"

datalist = [2,4,8,16,32,64,128,256] #임의의 숫자 리스트 정의

i=5  
for  data  in  datalist:
    ws.cell(row = i, column=1).value = data  #A열(Column=1)에 행을 바꾸면서 입력
    i=i+1  

wb.save("result2.xlsx") #엑셀 파일 저장

